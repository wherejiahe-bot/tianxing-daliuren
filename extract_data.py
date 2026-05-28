#!/usr/bin/env python
"""提取 天星2.0.xlsx 的全部数据为 JSON 结构"""
import json, openpyxl
from openpyxl.utils import get_column_letter as col_letter

SRC = 'F:/易学赚钱/大六壬/天星2.0.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['天星1.0']

MAX_R, MAX_C = ws.max_row, ws.max_column

# 1. 提取所有有值单元格
cells = {}
for r in range(1, MAX_R + 1):
    for c in range(1, MAX_C + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() not in ('#N/A', ''):
            cells.setdefault(r, {})[col_letter(c)] = str(v)


# 2. 提取合并单元格
merged = [str(m) for m in ws.merged_cells.ranges]


# 3. 提取行1-3的排盘输入区
header_area = {}
for r in range(1, 4):
    header_area[r] = cells.get(r, {})


# 4. 提取鲁班法/神煞数据 (Rows 4-20)
shensha_area = {}
for r in range(4, 21):
    shensha_area[r] = cells.get(r, {})

# 5. 旬甲表 (Rows 21-45)
xunjia_table = {}
for r in range(21, 46):
    xunjia_table[r] = cells.get(r, {})

# 6. 解神表 (Rows 48-60)
jieshen_table = {}
for r in range(48, 61):
    jieshen_table[r] = cells.get(r, {})

# 7. 小六壬表 (Rows 61-70)
xiaoliuren_table = {}
for r in range(61, 71):
    xiaoliuren_table[r] = cells.get(r, {})

# 8. 天地人表 (Rows 71-76)
tiandiren_table = {}
for r in range(71, 77):
    tiandiren_table[r] = cells.get(r, {})

# 9. 天星数据表 (Rows 64-120+) - 按年索引
yearly_data = {}
for r in range(64, min(MAX_R + 1, 121)):
    d = cells.get(r, {})
    if d:
        # 提取关键年份信息
        year_info = {}
        for k in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O',
                  'XY','XZ','YA','YB','YD','ZA','ZB','ZC']:
            if k in d:
                year_info[k] = d[k]
        yearly_data[r] = year_info

# 10. 右侧详表 (Col DK+) - 大六壬天将等
right_table = {}
for r in range(4, 21):
    d = {}
    for c in range(100, min(MAX_C + 1, 250)):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() not in ('#N/A', ''):
            d[col_letter(c)] = str(v)[:30]
    if d:
        right_table[r] = d

# 11. 行64+的完整右半数据
right_yearly = {}
for r in range(64, min(MAX_R + 1, 121)):
    d = {}
    for c in range(70, min(MAX_C + 1, 300)):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() not in ('#N/A', ''):
            d[col_letter(c)] = str(v)[:30]
    if d:
        right_yearly[r] = d

output = {
    'sheet_name': '天星1.0',
    'rows': MAX_R,
    'cols': MAX_C,
    'merged_cells': merged,
    'header_area': header_area,
    'shensha_area': shensha_area,
    'xunjia_table': xunjia_table,
    'jieshen_table': jieshen_table,
    'xiaoliuren_table': xiaoliuren_table,
    'tiandiren_table': tiandiren_table,
    'yearly_data': yearly_data,
    'right_table': right_table,
    'right_yearly': right_yearly,
    'all_cells': cells,
}

with open('G:/WorkBuddy/码农龙虾/tianxing-daliuren/export_data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=1)

print(f'Extracted: {len(cells)} rows with data, {len(merged)} merged cells')
print(f'Yearly data rows: {len(yearly_data)}')
