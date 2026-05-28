#!/usr/bin/env python
"""完整重写：提取天星2.0数据 → 生成 HTML"""
import json, openpyxl
from openpyxl.utils import get_column_letter as cl

SRC = 'F:/易学赚钱/大六壬/天星2.0.xlsx'

# 1. 先用 data_only=True 提取数值，再用 formula 模式提取公式
wb_val = openpyxl.load_workbook(SRC, data_only=True)
wb_fml = openpyxl.load_workbook(SRC, data_only=False)

def get_val(ws, r, c):
    """优先取缓存值，没有则取公式"""
    v = ws.cell(row=r, column=c).value
    if v is None:
        v = ''
    s = str(v).strip()
    return '' if s in ('None', '#N/A', '') else s

# ========== 天星1.0 sheet 的视觉分区 ==========
ws = wb_val['天星1.0']
wf = wb_fml['天星1.0']
MAX_R, MAX_C = ws.max_row, ws.max_column

# === A. 所有有值单元格（用于查表）===
all_cells_val = {}
all_cells_fml = {}
for r in range(1, MAX_R + 1):
    for c in range(1, MAX_C + 1):
        vv = get_val(ws, r, c)
        vf = get_val(wf, r, c)
        col = cl(c)
        if vv:
            all_cells_val.setdefault(r, {})[col] = vv
        if vf and vf != vv:
            all_cells_fml.setdefault(r, {})[col] = vf

def cv(r, c):
    """读取缓存值"""
    col = cl(c) if isinstance(c, int) else c
    if isinstance(c, int):
        col = cl(c)
    return all_cells_val.get(r, {}).get(col, '')

def cf(r, c):
    """读取公式"""
    col = cl(c) if isinstance(c, int) else c
    return all_cells_fml.get(r, {}).get(col, '')

# ========== 生成 HTML ==========
H = []
H.append("""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>天星2.0 · 大六壬排盘系统</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Microsoft YaHei','PingFang SC','Noto Sans SC',sans-serif;background:#0d1117;color:#c9d1d9;padding:16px}
.container{max-width:1500px;margin:0 auto}
h1{text-align:center;color:#ffd700;margin-bottom:4px;font-size:20px;letter-spacing:3px}
.subtitle{text-align:center;color:#8b949e;font-size:12px;margin-bottom:16px}
.card{background:#161b22;border:1px solid #30363d;border-radius:6px;padding:12px;margin-bottom:12px}
.card-title{color:#ffd700;font-size:13px;font-weight:bold;margin-bottom:8px;padding-bottom:5px;border-bottom:1px solid #30363d}
table{width:100%;border-collapse:collapse;font-size:12px}
th,td{border:1px solid #30363d;padding:3px 6px;text-align:center;white-space:nowrap}
th{background:#1c2333;color:#ffd700;font-weight:600}
td{background:#0d1117}
tr:hover td{background:#1c2333}
.table-wrap{overflow-x:auto;overflow-y:auto;max-height:480px}
.header-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(90px,1fr));gap:6px}
.header-item{background:#0d1117;border:1px solid #30363d;border-radius:4px;padding:5px 8px;text-align:center}
.header-item .label{color:#8b949e;font-size:10px}
.header-item .value{color:#ffd700;font-size:14px;font-weight:bold;margin-top:2px}
.header-item .value.sub{color:#58a6ff;font-size:12px}
.pan-grid{display:grid;gap:8px}
.pan-row{display:flex;gap:6px;flex-wrap:wrap}
.pan-cell{background:#0d1117;border:1px solid #30363d;border-radius:4px;padding:5px 8px;min-width:60px;flex:1;text-align:center}
.pan-cell .label{color:#666;font-size:10px}
.pan-cell .val{color:#c9d1d9;font-size:13px;font-weight:bold}
.val.gold{color:#ffd700}.val.green{color:#7ee787}.val.red{color:#ff7b72}.val.blue{color:#79c0ff}.val.purple{color:#d2a8ff}
.xlr-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:8px}
.xlr-card{background:#0d1117;border:1px solid #30363d;border-radius:6px;padding:8px;text-align:center}
.xlr-name{font-size:16px;font-weight:bold}
.xlr-sub{font-size:11px;color:#8b949e;margin-top:3px}
.daan .xlr-name{color:#7ee787}
.liulian .xlr-name{color:#ffd700}
.suxi .xlr-name{color:#79c0ff}
.chikou .xlr-name{color:#ff7b72}
.xiaoji .xlr-name{color:#d2a8ff}
.kongwang .xlr-name{color:#666}
.shensha-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:3px}
.shensha-item{background:#0d1117;border:1px solid #21262d;border-radius:3px;padding:2px 4px;font-size:10px;text-align:center}
.shensha-item .name{color:#8b949e;font-size:9px}
.shensha-item .val{color:#c9d1d9}
.note{color:#8b949e;font-size:10px;padding:4px 8px;line-height:1.5}
.col-year{color:#ffd700;font-weight:bold}
.col-gan{color:#79c0ff}
.col-zhi{color:#7ee787}
.col-nayin{color:#d2a8ff}
.col-wang{color:#ff7b72}
@media(max-width:768px){.xlr-grid{grid-template-columns:repeat(2,1fr)}.shensha-grid{grid-template-columns:repeat(3,1fr)}}
</style>
</head>
<body>
<div class="container">
<h1>✦ 天星2.0 · 大六壬排盘 ✦</h1>
<p class="subtitle">九天玄数 | 大六壬 | 小六壬 | 神煞集成 · 数据开放查询</p>
""")

# ========== 区块1: 四柱信息 ==========
H.append('<div class="card"><div class="card-title">📐 四柱信息</div><div class="header-grid">')

# 从 Row1 获取标签，Row2-3 获取值
labels = {'E':'年','F':'月','G':'日','H':'时','I':'月将','J':'性别','K':'年份','L':'本命','M':'行年'}
for col, label in labels.items():
    v2 = cv(2, col)
    v3 = cv(3, col)
    H.append(f'<div class="header-item"><div class="label">{label}</div>')
    if v2:
        H.append(f'<div class="value">{v2}</div>')
    if v3:
        H.append(f'<div class="value sub">{v3}</div>')
    H.append('</div>')

# 旬/空/月破
for col, label in [('DU','旬'),('DV','空'),('DP','月破')]:
    v = cv(1, col)
    if v:
        H.append(f'<div class="header-item"><div class="label">{label}</div><div class="value">{v}</div></div>')

H.append('</div>')

# 排盘说明
note = cv(1, 'AI')
if note:
    H.append(f'<div class="note">{note.replace(chr(10),"<br>")}</div>')
H.append('</div>')

# ========== 区块2: 日干神煞 ==========
H.append('<div class="card"><div class="card-title">🔮 日干神煞</div><div class="shensha-grid">')

# 读取 U-AF 列的神煞数据 (Row 1 = 标题, Row 2+ = 值)
shensha_label_row = 1
for r in [1, 2, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
    for col in ['U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF']:
        v = cv(r, col)
        if v and v in ('#N/A','N/A','',): 
            continue
        if v:
            label = cv(1, col) or col
            H.append(f'<div class="shensha-item"><span class="name">{label}</span><br><span class="val">{v}</span></div>')

H.append('</div></div>')

# ========== 区块3: 小六壬 ==========
H.append('<div class="card"><div class="card-title">☯ 小六壬掌诀</div><div class="xlr-grid">')

xlr_names = [('大安','daan'),('留连','liulian'),('速喜','suxi'),('赤口','chikou'),('小吉','xiaoji'),('空亡','kongwang')]
xlr_detail = {}
for r in range(65, 71):
    name = cv(r, 'F')
    if name:
        xlr_detail[name] = {
            'num': cv(r, 'E'),
            'shu': cv(r, 'G'),
            'di': cv(r, 'H'),
            'tian': cv(r, 'I'),
            'shen': cv(r, 'N'),
            'ji': cv(r, 'K'),
        }

# Also check rows 78-93 for detail versions
for r in range(78, 94):
    name = cv(r, 'H')
    if name and name in xlr_detail:
        xlr_detail[name].setdefault('detail_positions', []).append(cv(r, 'L'))

for name, css_class in xlr_names:
    d = xlr_detail.get(name, {})
    H.append(f'<div class="xlr-card {css_class}">')
    H.append(f'<div class="xlr-name">{name}</div>')
    if d.get('num'): H.append(f'<div class="xlr-sub">位: {d["num"]}</div>')
    if d.get('di'): H.append(f'<div class="xlr-sub">地: {d["di"]}</div>')
    if d.get('tian'): H.append(f'<div class="xlr-sub">天: {d["tian"]}</div>')
    if d.get('shen'): H.append(f'<div class="xlr-sub">神: {d["shen"]}</div>')
    if d.get('ji'): H.append(f'<div class="xlr-sub">将: {d["ji"]}</div>')
    if d.get('detail_positions'):
        H.append(f'<div class="xlr-sub" style="color:#666;">位: {",".join(d["detail_positions"])}</div>')
    H.append('</div>')

H.append('</div></div>')

# ========== 区块4: 天地人 + 八卦 ==========
H.append('<div class="card"><div class="card-title">🌌 天地人 · 八卦 · 玄数</div>')

# Rows 62-63: 卦象
for r in [62, 63]:
    v = cv(r, 'V')
    w = cv(r, 'W')
    if v:
        H.append(f'<div class="pan-row">')
        H.append(f'<div class="pan-cell"><div class="label">卦</div><div class="val blue">{v}</div></div>')
        if w: H.append(f'<div class="pan-cell"><div class="label">数</div><div class="val green">{w}</div></div>')
        H.append('</div>')

# Rows 71-76: 天地人位
for r in range(71, 77):
    a = cv(r, 'A')  # 天地人
    n = cv(r, 'N')  # 神
    if a or n:
        H.append('<div class="pan-row">')
        if a: H.append(f'<div class="pan-cell"><div class="label">位</div><div class="val gold">{a}</div></div>')
        if n: H.append(f'<div class="pan-cell"><div class="label">神</div><div class="val purple">{n}</div></div>')
        H.append('</div>')

H.append('</div>')

# ========== 区块5: 年表 ==========
H.append('<div class="card"><div class="card-title">📅 六十甲子 · 天星年表</div>')
H.append('<div class="table-wrap"><table>')
H.append('<thead><tr><th>年</th><th>干支</th><th>月柱</th><th>日柱</th><th>时柱</th><th>纳音</th><th>旺/休/囚</th><th>神</th><th>将</th><th>小六壬</th><th>地</th><th>天星</th></tr></thead><tbody>')

for r in range(64, 121):
    year = cv(r, 'XY')
    if not year:
        continue
    hlist = [
        f'<td class="col-year">{year}</td>',
        f'<td class="col-gan">{cv(r,"XZ")}</td>',
        f'<td>{cv(r,"YA")}</td>',
        f'<td>{cv(r,"YB")}</td>',
        f'<td>{cv(r,"YC")}</td>',
        f'<td class="col-nayin">{cv(r,"F") or cv(r,"FM") or ""}</td>',
        f'<td class="col-wang">{cv(r,"C") or ""}</td>',
        f'<td>{cv(r,"K") or cv(r,"EK") or ""}</td>',
        f'<td>{cv(r,"N") or ""}</td>',
        f'<td>{cv(r,"H") or cv(r,"EJ") or cv(r,"F") or ""}</td>',
        f'<td>{cv(r,"L") or cv(r,"ZB") or ""}</td>',
        f'<td class="col-gan">{cv(r,"ZC") or ""}</td>',
    ]
    H.append('<tr>' + ''.join(hlist) + '</tr>')

H.append('</tbody></table></div></div>')

# ========== 区块6: 右侧详表 ==========
H.append('<div class="card"><div class="card-title">📜 大六壬 · 天将神煞全景</div>')
H.append('<div class="table-wrap"><table><thead><tr>')

key_cols = ['DK','DL','DM','DN','DO','DP','DQ','DR','DS','DT','DU','DV','DW',
            'EJ','EK','EL','EM','EN','EO','EP','EQ','ER','ES','ET','EU',
            'EW','EX','EY','EZ','FA','FB','FC','FD','FE','FF','FG','FH','FI','FJ']

# Row 4 = headers
for c in key_cols:
    hdr = cv(4, c) or c
    H.append(f'<th>{hdr}</th>')
H.append('</tr></thead><tbody>')

for r in range(5, 21):
    vals = {c: cv(r, c) for c in key_cols}
    if not any(vals.values()):
        continue
    H.append('<tr>')
    for c in key_cols:
        H.append(f'<td>{vals[c] or ""}</td>')
    H.append('</tr>')

H.append('</tbody></table></div></div>')

# ========== 页脚 ==========
H.append("""<div style="text-align:center;color:#484f58;font-size:11px;padding:16px 0;">
天星2.0 · 大六壬排盘系统 · GitHub Pages 版 · By 霎哈嘉瑜伽修习者
</div>
</div>
</body>
</html>""")

output = 'G:/WorkBuddy/码农龙虾/tianxing-daliuren/index.html'
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(H))

print(f'✓ HTML generated: {output}')
print(f'  Size: {sum(len(h) for h in H)} bytes')
